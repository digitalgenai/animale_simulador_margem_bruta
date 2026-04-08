from __future__ import annotations

import tempfile
from pathlib import Path
from openpyxl import load_workbook
import logging
import re
import threading
from datetime import datetime
from typing import Any, Dict, List, Tuple

import pandas as pd

from dash import Dash, html, dcc, Input, Output, State, ctx, no_update
import dash_bootstrap_components as dbc
import dash_ag_grid as dag

from core.config import (
    BASE_SIMULADOR_PATH,
    COLUNA_AGREGACAO_PRINCIPAL,
)
from core.data_loader import load_base_data, get_month_context
from core.view_builders import (
    compute_summary,
    build_tab1_rows,
    build_tab2_rows,
    build_tab3_rows,
    build_history_payload,
)
from core.calculations import calcular_custo_necessario, calcular_margem_real_percentual

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("simulador_web")

# =============================================================================
# Mapeamento Tipo de Embalagem (fonte: tipo_embalagem.xlsx)
# =============================================================================
_TIPO_EMBAL_MAP: Dict[str, str] = {}
_TIPO_EMBAL_OPCOES: List[str] = ["[TODAS]"]
try:
    _df_embal = pd.read_excel(
        r"C:\Users\ranyer.paiva\Downloads\tipo_embalagem.xlsx",
        sheet_name="Base",
        dtype=str,
    )
    _df_embal.columns = [c.strip() for c in _df_embal.columns]
    _bc_col = next(c for c in _df_embal.columns if "barras" in c.lower())
    _tp_col = next(c for c in _df_embal.columns if "embalagem" in c.lower())
    _df_embal[_bc_col] = _df_embal[_bc_col].str.strip()
    _df_embal[_tp_col] = _df_embal[_tp_col].str.strip()
    _TIPO_EMBAL_MAP = dict(zip(_df_embal[_bc_col], _df_embal[_tp_col]))
    _tipos_unicos = sorted(_df_embal[_tp_col].dropna().unique().tolist())
    _TIPO_EMBAL_OPCOES = ["[TODAS]"] + _tipos_unicos
    del _df_embal, _bc_col, _tp_col, _tipos_unicos
except Exception as _e:
    logger.warning("Não foi possível carregar tipo_embalagem.xlsx: %s", _e)

# =============================================================================
# Cache de datasets por Mês/Ano (evita reload em toda interação)
# TTL de 30 minutos: garante que dados do mês vigente sejam sempre frescos
# =============================================================================
_DATA_LOCK = threading.Lock()
_CACHE_TTL_SECONDS = 1800  # 30 minutos
_DATA_CACHE: Dict[str, Tuple[Tuple, datetime]] = {}


def _parse_ym_safe(s: str | None) -> Tuple[int | None, int | None]:
    if not s:
        return None, None
    m = re.match(r"^(\d{4})[_-](\d{2})$", str(s).strip())
    if not m:
        return None, None
    return int(m.group(1)), int(m.group(2))


def _get_data_for_mes_ref(mes_ref_safe: str | None, force_reload: bool = False, mes_inicio_safe: str | None = None):
    key = f"{mes_ref_safe or '__DEFAULT__'}|{mes_inicio_safe or ''}"

    with _DATA_LOCK:
        if force_reload:
            _DATA_CACHE.pop(key, None)

        cached = _DATA_CACHE.get(key)
        if cached:
            result, loaded_at = cached
            if (datetime.now() - loaded_at).total_seconds() < _CACHE_TTL_SECONDS:
                return result
            _DATA_CACHE.pop(key, None)

        # carrega ainda sob lock (evita corrida em callbacks paralelos)
        y_ini, m_ini = _parse_ym_safe(mes_inicio_safe) if mes_inicio_safe else (None, None)

        if mes_ref_safe is None or mes_ref_safe == "__DEFAULT__":
            df_base, bench_ano, bench_6m, bench_3m, lista_fornecedores, lista_categorias_global = load_base_data(
                ref_start_year=y_ini, ref_start_month_num=m_ini
            )
            month_ctx = get_month_context()
        else:
            y, m = _parse_ym_safe(mes_ref_safe)
            if y is None or m is None:
                df_base, bench_ano, bench_6m, bench_3m, lista_fornecedores, lista_categorias_global = load_base_data(
                    ref_start_year=y_ini, ref_start_month_num=m_ini
                )
                month_ctx = get_month_context()
            else:
                df_base, bench_ano, bench_6m, bench_3m, lista_fornecedores, lista_categorias_global = load_base_data(
                    ref_year=y, ref_month=m, ref_start_year=y_ini, ref_start_month_num=m_ini
                )
                month_ctx = get_month_context()

        # Enriquecer com Tipo de Embalagem (join pelo Cod_Barras)
        if _TIPO_EMBAL_MAP and isinstance(df_base, pd.DataFrame) and not df_base.empty:
            df_base["Tipo_Embalagem"] = (
                df_base["Cod_Barras"].map(_TIPO_EMBAL_MAP).fillna("NÃO TEM")
            )

        result = (df_base, bench_ano, bench_6m, bench_3m, lista_fornecedores, lista_categorias_global, month_ctx)
        _DATA_CACHE[key] = (result, datetime.now())
        return result


def _get_available_safe() -> list:
    with _DATA_LOCK:
        for cached_val in _DATA_CACHE.values():
            result, _ = cached_val
            ctx_candidate = result[6] if result and len(result) > 6 else {}
            avail = (ctx_candidate or {}).get("available_labels_safe")
            if avail:
                return list(avail)
    return list(month_ctx0.get("available_labels_safe") or [])


def _resolve_mes_ref(periodo_tipo: str | None, mes_ref: str | None, mes_fim: str | None) -> str | None:
    """Retorna o mês final efetivo: quando Personalizado usa mes_fim, senão usa mes_ref."""
    if periodo_tipo == "personalizado" and mes_fim:
        return mes_fim
    if periodo_tipo in ("mes_anterior", "ontem"):
        available_safe = _get_available_safe()
        if not mes_ref or not available_safe:
            return mes_ref
        try:
            idx = list(available_safe).index(mes_ref)
            start_idx = max(0, idx - 1)
            return available_safe[start_idx]
        except ValueError:
            return mes_ref
    return mes_ref


def _resolve_mes_inicio(periodo_tipo: str | None, mes_ref_safe: str | None, mes_inicio_safe: str | None, available_safe: list) -> str | None:
    """Deriva mes_inicio_safe a partir do tipo de período selecionado."""
    if not periodo_tipo or periodo_tipo in ("mes_unico", "hoje", "mes_anterior", "ontem"):
        return None
    if periodo_tipo == "personalizado":
        return mes_inicio_safe
    n_map = {"ultimos_3": 3, "ultimos_6": 6, "ultimos_12": 12}
    n = n_map.get(str(periodo_tipo), 1)
    if not mes_ref_safe or not available_safe:
        return None
    try:
        idx = list(available_safe).index(mes_ref_safe)
    except ValueError:
        return None
    start_idx = max(0, idx - (n - 1))
    return available_safe[start_idx]


# --- Carga default (para montar layout inicial) ---
try:
    df_base0, bench_ano0, bench_6m0, bench_3m0, lista_fornecedores0, lista_categorias0 = load_base_data()
    month_ctx0 = get_month_context()
except Exception as e:
    logger.exception("Erro inicialização ao carregar '%s': %s", BASE_SIMULADOR_PATH, e)
    df_base0 = pd.DataFrame()
    bench_ano0, bench_6m0, bench_3m0 = {}, {}, {}
    lista_fornecedores0, lista_categorias0 = ["SEM DADOS"], []
    month_ctx0 = {}

# options do seletor Mês/Ano (range real do dataset)
_available_safe = month_ctx0.get("available_labels_safe") or []
_available_human = month_ctx0.get("available_labels_human") or []
MES_REF_OPTIONS = (
    [{"label": h, "value": s} for s, h in zip(_available_safe, _available_human)]
    if _available_safe and _available_human and len(_available_safe) == len(_available_human)
    else []
)

DEFAULT_MES_REF_SAFE = month_ctx0.get("ref_month_safe") if month_ctx0 else None
if _available_safe:
    DEFAULT_MES_REF_SAFE = _available_safe[-1]


def _safe_to_ddmmyyyy(safe: str | None) -> str:
    """Converte 'YYYY_MM' para '01/MM/YYYY' (texto de exibição para o input de data)."""
    if not safe:
        return ""
    m = re.match(r"^(\d{4})[_-](\d{2})$", str(safe).strip())
    if m:
        return f"01/{m.group(2)}/{m.group(1)}"
    return ""


def _parse_ddmmyyyy_to_safe(s: str | None) -> str | None:
    """Converte 'dd/mm/yyyy' para 'YYYY_MM'. Aceita também 'YYYY_MM' passado diretamente."""
    if not s:
        return None
    s = str(s).strip()
    # Formato dd/mm/yyyy
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        return f"{m.group(3)}_{m.group(2).zfill(2)}"
    # Já em formato YYYY_MM ou YYYY-MM (retrocompat)
    if re.match(r"^\d{4}[_-]\d{2}$", s):
        return s.replace("-", "_")
    return None


DEFAULT_MES_FIM_TEXT = _safe_to_ddmmyyyy(DEFAULT_MES_REF_SAFE)


# ---------- Helpers ----------
def _safe_float_percent(val: Any, default: float) -> float:
    """
    Recebe string tipo "30.0" ou "30,0" e devolve FRAÇÃO (0.30).
    """
    try:
        if val is None:
            return default
        s = str(val).replace(",", ".").strip()
        return float(s) / 100.0
    except Exception:
        return default


def _filter_tab12(df_base: pd.DataFrame, forn: str, fab: str, cat: str, tipo_embal: str = "[TODAS]") -> pd.DataFrame:
    if df_base is None or df_base.empty:
        return df_base.iloc[0:0].copy() if isinstance(df_base, pd.DataFrame) else pd.DataFrame()

    if not forn or forn == "[TODOS]":
        df_temp = df_base.copy()
    else:
        df_temp = df_base[df_base[COLUNA_AGREGACAO_PRINCIPAL] == forn]

    if fab and fab != "[TODOS]":
        df_temp = df_temp[df_temp["Fabricante"] == fab]

    if cat and cat != "[TODAS]":
        df_temp = df_temp[df_temp["Area"] == cat]

    if tipo_embal and tipo_embal != "[TODAS]" and "Tipo_Embalagem" in df_temp.columns:
        df_temp = df_temp[df_temp["Tipo_Embalagem"] == tipo_embal]

    abc_map = {"A": 0, "B": 1, "C": 2}
    abc_order = df_temp["Curva_ABC"].map(abc_map).fillna(3)
    df_temp = df_temp.assign(ABC_Order=abc_order).sort_values(
        ["ABC_Order", "Fat_Ref"], ascending=[True, False]
    )
    return df_temp


def _filter_tab3(df_base: pd.DataFrame, cat_t3: str, forn_t3: str) -> pd.DataFrame:
    if df_base is None or df_base.empty or not cat_t3:
        return df_base.iloc[0:0].copy() if isinstance(df_base, pd.DataFrame) else pd.DataFrame()

    df_temp = df_base[df_base["Area"] == cat_t3]
    if forn_t3 and forn_t3 != "[TODOS]":
        df_temp = df_temp[df_temp["Fornecedor"] == forn_t3]
    return df_temp


def _get_fab_cat_options_for_supplier(df_base: pd.DataFrame, forn: str) -> Tuple[List[str], List[str]]:
    if df_base is None or df_base.empty or not forn or forn == "[TODOS]":
        return ["[TODOS]"], ["[TODAS]"]

    df_forn = df_base[df_base[COLUNA_AGREGACAO_PRINCIPAL] == forn]

    lista_fab = sorted(df_forn["Fabricante"].unique().tolist())
    lista_fab.insert(0, "[TODOS]")

    hidden_cats = {"outros", "outro", "desconhecidos", "desconhecido"}
    lista_cat = sorted(
        [
            x
            for x in df_forn["Area"].astype(str).unique().tolist()
            if str(x).strip().lower() not in hidden_cats
        ]
    )
    lista_cat.insert(0, "[TODAS]")

    return lista_fab, lista_cat


def _get_supplier_options_for_category(df_base: pd.DataFrame, cat_t3: str) -> List[str]:
    if df_base is None or df_base.empty or not cat_t3:
        return ["[TODOS]"]

    if str(cat_t3).strip().lower() in {"outros", "outro", "desconhecidos", "desconhecido"}:
        return ["[TODOS]"]

    df_cat = df_base[df_base["Area"] == cat_t3]
    rank_forn_cat = df_cat.groupby("Fornecedor")["Fat_Ref"].sum().sort_values(ascending=False)
    lista_forn_cat = rank_forn_cat.index.tolist()
    lista_forn_cat.insert(0, "[TODOS]")
    return lista_forn_cat


def _apply_row_class_rules() -> Dict[str, Any]:
    return {
        "rowClassRules": {
            "row-neg": "params.data.__is_neg === true",
            "row-yellow": "params.data.__is_yellow === true",
        }
    }


def _is_hidden_category(val: Any) -> bool:
    s = str(val or "").strip().lower()
    return s in {"outros", "outro", "desconhecidos", "desconhecido"}


def _format_currency_br(v: float) -> str:
    try:
        s = f"{float(v):,.2f}"
        return "R$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"


def _parse_br_number_like_excel(val):
    if val is None:
        return None

    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)

    s = str(val).strip()
    if not s:
        return None

    is_percent = "%" in s

    # remove símbolos monetários e espaços
    s = (
        s.replace("R$", "")
         .replace("$", "")
         .replace("%", "")
         .replace("\u00a0", "")
         .replace(" ", "")
         .strip()
    )

    if not s:
        return None

    # caso tenha negativo entre parênteses: (1.234,56)
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1].strip()

    # se tem . e ,:
    # - se a última vírgula vem depois do último ponto => padrão BR (1.234,56)
    # - senão => padrão US (1,234.56)
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            # BR
            s = s.replace(".", "").replace(",", ".")
        else:
            # US
            s = s.replace(",", "")
    elif "," in s:
        # só vírgula -> assume decimal BR
        s = s.replace(".", "").replace(",", ".")
    else:
        # só ponto ou inteiro -> mantém
        pass

    try:
        num = float(s)
        if neg:
            num = -num
        if is_percent:
            num /= 100.0
        return num
    except Exception:
        return None


def _apply_excel_formats(ws):
    currency_cols = {
        "Preço Atual",
        "Custo",
        "Marg R$",
        "Sim Preço",
        "Sim Custo",
        "PETZ",
        "PROCAMPO",
        "Sim Preço (Conc)",
        "Fat Ref",
        "Margem Ref R$",
        "Faturamento",
    }

    percent_cols = {
        "Marg %",
        "Sim Marg",
        "Marg Atual %",
        "Dif % (Menor)",
        "Dif % (menor preço)",
        "DELTA ALVO %",
        "Sim Margem (Result)",
        "Margem Ref %",
        "Fat_Acum_Pct",
    }

    integer_cols = {"Qtd Ref"}

    fmt_currency_br = 'R$ #,##0.00;[Red]-R$ #,##0.00'
    fmt_percent_2 = '0.00%'
    fmt_integer = '0'

    header_map = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value).strip() if cell.value is not None else ""
        header_map[header] = col_idx

    for header, col_idx in header_map.items():
        for row_idx in range(2, ws.max_row + 1):
            c = ws.cell(row=row_idx, column=col_idx)

            if header in currency_cols and isinstance(c.value, (int, float)):
                c.number_format = fmt_currency_br
            elif header in percent_cols and isinstance(c.value, (int, float)):
                c.number_format = fmt_percent_2
            elif header in integer_cols and isinstance(c.value, (int, float)):
                c.number_format = fmt_integer


def _format_kpi(label: str, value: str, color: str | None = None):
    style = {"display": "inline-block", "marginRight": "18px"}
    vstyle = {"fontWeight": "700"}
    if color:
        vstyle["color"] = color
    return html.Span(
        [
            html.Span(label, className="kpi-label"),
            html.Span(value, className="kpi-value", style=vstyle),
        ],
        style=style,
    )


def _breakdown_component(breakdown: List[Dict[str, Any]], month_ctx: Dict[str, Any] | None = None):
    lab = _closed_month_label(month_ctx)

    header = html.Thead(
        html.Tr(
            [
                html.Th("CATEGORIA"),
                html.Th(f"FAT({lab})"),
                html.Th(f"MG({lab})"),
                html.Th("GL(Year)"),
            ]
        )
    )
    body_rows = []
    for b in breakdown:
        cat = str(b.get("categoria", ""))
        if _is_hidden_category(cat):
            continue

        body_rows.append(
            html.Tr(
                [
                    html.Td(cat[:15]),
                    html.Td(_format_currency_br(b["fat"])),
                    html.Td(f"{b['marg_perc']:.2%}"),
                    html.Td(f"{b['bench_ano']:.2%}"),
                ]
            )
        )
    body = html.Tbody(body_rows)
    return dbc.Table([header, body], bordered=True, size="sm", className="breakdown-table")


def _history_component(hist: Dict[str, Any], suffix: str, month_ctx: Dict[str, Any] | None = None):
    lab = _closed_month_label(month_ctx)

    produto_val = hist.get("produto")
    produto_exib = produto_val if str(produto_val or "").strip() else "-"
    cod_barras_val = hist.get("cod_barras", "-")
    cod_barras_exib = cod_barras_val if str(cod_barras_val or "").strip() else "-"

    return html.Div(
        [
            html.Div("Detalhes (Inteligência Temporal):", style={"fontWeight": "700", "color": "navy"}),
            html.Div("Clique em um produto na tabela", style={"fontSize": "12px", "color": "#666", "marginBottom": "6px"}),

            html.Div(
                [
                    html.Span("Cod. Barras: ", style={"width": "90px", "display": "inline-block"}),
                    html.Span(cod_barras_exib, style={"fontFamily": "monospace"}),
                ]
            ),
            html.Div(
                [
                    html.Span("Produto: ", style={"width": "90px", "display": "inline-block"}),
                    html.Span(produto_exib, style={"fontStyle": "italic"}),
                ]
            ),
            html.Div(
                [
                    html.Span("Méd 6M: ", style={"width": "70px", "display": "inline-block"}),
                    html.Span(hist.get("hist_6m", "-"), style={"fontWeight": "700"}),
                    html.Span("  "),
                    html.Span("Méd 3M: ", style={"width": "70px", "display": "inline-block", "marginLeft": "10px"}),
                    html.Span(hist.get("hist_3m", "-"), style={"fontWeight": "700"}),
                ]
            ),
            html.Div(
                [
                    html.Span(f"Venda {lab}: ", style={"width": "90px", "display": "inline-block"}),
                    html.Span(hist.get("hist_ref", "-"), style={"fontWeight": "700", "color": "blue"}),
                    html.Span("  "),
                    html.Span("Pico: ", style={"width": "50px", "display": "inline-block", "marginLeft": "10px"}),
                    html.Span(hist.get("hist_pico", "-"), style={"fontWeight": "700", "color": "green"}),
                ]
            ),
        ],
        className="history-box",
        id=f"hist-box-{suffix}",
    )

def _get_last_closed_month_ts(month_ctx: Dict[str, Any] | None) -> pd.Timestamp | None:
    ctx2 = month_ctx or {}

    safe = ctx2.get("ref_month_safe") or ctx2.get("closed_month_safe")
    y, m = _parse_ym_safe(str(safe)) if safe else (None, None)
    if not y or not m:
        return None

    ts = pd.Timestamp(year=y, month=m, day=1)
    return ts - pd.DateOffset(months=1)

def _row_from_event(cell_event: dict, rowData: list[dict] | None):
    if not cell_event or not rowData:
        return None

    row_id = cell_event.get("rowId")
    if row_id is not None:
        for r in rowData:
            if str(r.get("id")) == str(row_id):
                return r

    idx = cell_event.get("rowIndex")
    if isinstance(idx, int) and 0 <= idx < len(rowData):
        return rowData[idx]

    return None


_PT_ABBR = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def _get_ref_month_ts(month_ctx: Dict[str, Any] | None) -> pd.Timestamp | None:
    ctx2 = month_ctx or {}
    safe = ctx2.get("ref_month_safe") or ctx2.get("closed_month_safe")
    y, m = _parse_ym_safe(str(safe)) if safe else (None, None)
    if not y or not m:
        return None
    return pd.Timestamp(year=y, month=m, day=1)


def _closed_month_label(month_ctx: Dict[str, Any] | None) -> str:
    ctx = month_ctx or {}
    ts_end = _get_ref_month_ts(ctx)

    ref_start_safe = ctx.get("ref_start_month_safe")
    if ref_start_safe:
        y_s, m_s = _parse_ym_safe(str(ref_start_safe))
        if y_s and m_s:
            ts_start = pd.Timestamp(year=y_s, month=m_s, day=1)
            if isinstance(ts_end, pd.Timestamp) and ts_start != ts_end:
                start_label = f"{_PT_ABBR[ts_start.month - 1]}/{ts_start.year}"
                end_label = f"{_PT_ABBR[ts_end.month - 1]}/{ts_end.year}"
                return f"{start_label}-{end_label}"

    if isinstance(ts_end, pd.Timestamp):
        return f"{_PT_ABBR[ts_end.month - 1]}/{ts_end.year}"
    return "Mês"


def _visible_fields_from_column_state(column_state):
    """
    Retorna lista de fields/colId visíveis no grid (hide != True).
    """
    if not isinstance(column_state, list) or not column_state:
        return []

    out = []
    for c in column_state:
        if not isinstance(c, dict):
            continue
        if c.get("hide") is True:
            continue
        # dash-ag-grid costuma usar colId; em muitos casos é igual ao field
        fid = c.get("colId") or c.get("field")
        if fid:
            out.append(str(fid))
    # remove duplicados preservando ordem
    seen = set()
    final = []
    for x in out:
        if x not in seen:
            seen.add(x)
            final.append(x)
    return final


# ---------- Dash app ----------
app = Dash(
    __name__,
    external_stylesheets=[dbc.themes.BOOTSTRAP],
    suppress_callback_exceptions=True,
    title="Simulador v76.8 - Web",
)
app.index_string = """
<!DOCTYPE html>
<html lang="pt-BR">
    <head>
        {%metas%}
        <meta charset="utf-8" />
        <meta http-equiv="Content-Language" content="pt-BR" />
        <meta name="language" content="pt-BR" />
        <meta name="viewport" content="width=device-width, initial-scale=1" />
        <title>{%title%}</title>
        {%favicon%}
        {%css%}
    </head>
    <body>
        {%app_entry%}
        <footer>
            {%config%}
            {%scripts%}
            {%renderer%}
        </footer>
    </body>
</html>
"""
server = app.server

store_sim_default = {"manual": {}, "conc": {}}

# ColumnDefs
coldefs_t1 = [
    {"headerName": "Cod. Barras", "field": "Cod_Barras", "width": 130},
    {"headerName": "Produto", "field": "Produto", "width": 260},
    {"headerName": "ABC", "field": "ABC", "width": 70},
    {"headerName": "Categ", "field": "Categ", "width": 140},
    {"headerName": "Qtd Ref", "field": "Qtd Ref", "width": 95},
    {"headerName": "Faturamento", "field": "Faturamento", "width": 130},
    {"headerName": "Fat. Acum %", "field": "Fat_Acum_Pct", "width": 110},
    {"headerName": "Preço Atual", "field": "Preço Atual", "width": 110},
    {"headerName": "Custo", "field": "Custo", "width": 110},
    {"headerName": "Marg R$", "field": "Marg R$", "width": 110},
    {"headerName": "Marg %", "field": "Marg %", "width": 95},
    {"headerName": "PETZ", "field": "PETZ", "width": 110},
    {"headerName": "PROCAMPO", "field": "PROCAMPO", "width": 110},
    {"headerName": "Dif % (menor preço)", "field": "Dif % (Menor)", "width": 120},
    {"headerName": "Menor preço", "field": "Sim Preço", "width": 110},
    {"headerName": "Sim Marg", "field": "Sim Marg", "width": 95},
    {"headerName": "Sim Custo", "field": "Sim Custo", "width": 120},
]

coldefs_t2 = [
    {"headerName": "Cod. Barras", "field": "Cod_Barras", "width": 130},
    {"headerName": "Produto", "field": "Produto", "width": 260},
    {"headerName": "ABC", "field": "ABC", "width": 70},
    {"headerName": "Categ", "field": "Categ", "width": 140},
    {"headerName": "Qtd Ref", "field": "Qtd Ref", "width": 95},
    {"headerName": "Faturamento", "field": "Faturamento", "width": 130},
    {"headerName": "Fat. Acum %", "field": "Fat_Acum_Pct", "width": 110},
    {"headerName": "Preço Atual", "field": "Preço Atual", "width": 110},
    {"headerName": "Custo", "field": "Custo", "width": 110},
    {"headerName": "Marg Atual %", "field": "Marg Atual %", "width": 115},
    {"headerName": "PETZ", "field": "PETZ", "width": 110},
    {"headerName": "PROCAMPO", "field": "PROCAMPO", "width": 110},
    {"headerName": "Dif % (menor preço)", "field": "Dif % (menor preço)", "width": 130},
    {"headerName": "DELTA ALVO %", "field": "DELTA ALVO %", "width": 120},
    {"headerName": "Sim Preço (Conc)", "field": "Sim Preço (Conc)", "width": 130},
    {"headerName": "Sim Margem (Result)", "field": "Sim Margem (Result)", "width": 140},
]

coldefs_t3 = [
    {"headerName": "Fornecedor", "field": "Fornecedor", "width": 320},
    {"headerName": "Fat Ref", "field": "Fat Ref", "width": 160},
    {"headerName": "Fat. Acum %", "field": "Fat_Acum_Pct", "width": 120},
    {"headerName": "R$ Margem", "field": "Margem Ref R$", "width": 160},
    {"headerName": "% Margem", "field": "Margem Ref %", "width": 140},
]


def make_grid(grid_id: str, column_defs: List[Dict[str, Any]]) -> dag.AgGrid:
    return dag.AgGrid(
        id=grid_id,
        columnDefs=column_defs,
        rowData=[],
        getRowId="params.data.id",
        defaultColDef={
            "resizable": True,
            "sortable": True,
            "filter": True,
            "wrapHeaderText": True,
            "autoHeaderHeight": True,
        },
        dashGridOptions={
            "rowSelection": "single",
            "suppressRowClickSelection": False,
            "rowMultiSelectWithClick": False,
            "animateRows": True,
            **_apply_row_class_rules(),
        },
        className="ag-theme-alpine",
        style={"height": "520px", "width": "100%", "minWidth": "0"},
    )


def make_summary_block(suffix: str):
    return dbc.Card(
        dbc.CardBody(
            [
                html.Div(
                    [
                        _format_kpi("Fat. Total:", "-", None),
                        _format_kpi("Margem Média:", "-", "blue"),
                        html.Span("| ", style={"color": "#999"}),
                        _format_kpi("Total SKUs:", "-", None),
                        _format_kpi("A:", "-", "green"),
                        _format_kpi("B:", "-", "#bda404"),
                        _format_kpi("C:", "-", "red"),
                    ],
                    id=f"kpi-line-{suffix}",
                ),
              html.Div(
                    'Obs.: O faturamento total e a Margem Média incluem itens classificados como "Outros/Desconhecidos".',
                    className="small-muted",
                    style={"marginTop": "6px"},
                ),
                html.Hr(),
                dbc.Row(
                    [
                        dbc.Col(
                            [
                                html.Div("Top Categorias (Forn. vs Benchmarks):", style={"fontWeight": "700"}),
                                html.Div(id=f"breakdown-{suffix}", children=_breakdown_component([], month_ctx0)),
                            ],
                            md=7,
                        ),
                        dbc.Col(
                            _history_component(
                                {"produto": "-", "cod_barras": "-", "hist_6m": "-", "hist_3m": "-", "hist_ref": "-", "hist_pico": "-"},
                                suffix,
                                month_ctx0,
                            ),
                            md=5,
                        ),
                    ],
                    className="g-2",
                ),
            ]
        ),
        className="mb-2",
    )


# ---------- Layout ----------
app.layout = dbc.Container(
    fluid=True,
    children=[
        dcc.Store(id="store-sim", storage_type="session", data=store_sim_default),
        dcc.Store(id="store-selected", storage_type="session", data={"produto_key": None, "area": ""}),
        dcc.Download(id="download-excel"),
        dcc.Interval(id="interval-refresh", interval=_CACHE_TTL_SECONDS * 1000, n_intervals=0),

        dbc.Row(
            dbc.Col(
                dbc.Card(
                    dbc.CardBody(
                        [
                            html.Div([
                                # Linha 1: Principal
                                dbc.Row([
                                    dbc.Col([
                                        html.Div("Período de Análise", className="small text-muted fw-bold mb-1"),
                                        dcc.Dropdown(
                                            id="periodo_tipo",
                                            options=[
                                                {"label": "Últimos 3 meses", "value": "ultimos_3"},
                                                {"label": "Últimos 6 meses", "value": "ultimos_6"},
                                                {"label": "Últimos 12 meses", "value": "ultimos_12"},
                                                {"label": "Mês Atual", "value": "mes_unico"},
                                                {"label": "Mês Anterior", "value": "mes_anterior"},
                                                {"label": "Personalizado", "value": "personalizado"},
                                            ],
                                            value="mes_unico",
                                            clearable=False,
                                            searchable=True,
                                            className="shadow-sm"
                                        )
                                    ], md=4, lg=2, className="mb-3"),
                                    dbc.Col([
                                        html.Div("Mês/Ano Principal", className="small text-muted fw-bold mb-1"),
                                        dcc.Dropdown(
                                            id="mes_ref",
                                            options=MES_REF_OPTIONS,
                                            value=DEFAULT_MES_REF_SAFE,
                                            placeholder="Selecione...",
                                            clearable=False,
                                            searchable=True,
                                            className="shadow-sm"
                                        )
                                    ], id="col-mes-ref", md=4, lg=2, className="mb-3"),
                                    dbc.Col([
                                        html.Div("Data Início", className="small text-muted fw-bold mb-1"),
                                        dcc.Input(
                                            id="mes_inicio",
                                            type="text",
                                            value=None,
                                            placeholder="dd/mm/aaaa",
                                            debounce=True,
                                            className="form-control shadow-sm",
                                            style={"fontSize": "14px"},
                                        )
                                    ], id="col-mes-inicio", md=4, lg=2, className="mb-3", style={"display": "none"}),
                                    dbc.Col([
                                        html.Div("Data Fim", className="small text-muted fw-bold mb-1"),
                                        dcc.Input(
                                            id="mes_fim",
                                            type="text",
                                            value=DEFAULT_MES_FIM_TEXT,
                                            placeholder="dd/mm/aaaa",
                                            debounce=True,
                                            className="form-control shadow-sm",
                                            style={"fontSize": "14px"},
                                        )
                                    ], id="col-mes-fim", md=4, lg=2, className="mb-3", style={"display": "none"}),
                                    dbc.Col([
                                        dbc.Button("Atualizar Resumo", id="btn-atualizar", color="primary", className="w-100 shadow-sm fw-bold")
                                    ], md=2, lg=2, className="ms-md-auto mb-3 align-self-end"),
                                    dbc.Col([
                                        dbc.Button("Exportar Excel", id="btn-export", color="success", className="w-100 shadow-sm fw-bold")
                                    ], md=2, lg=2, className="mb-3 align-self-end"),
                                ], className="g-3 align-items-end"),
                                html.Hr(className="text-muted mt-0 mb-4"),

                                # Linha 2 e 3: Filtros e Metas
                                dbc.Row([
                                    dbc.Col([
                                        html.Div(f"{COLUNA_AGREGACAO_PRINCIPAL}", className="small text-muted fw-bold mb-1"),
                                        dcc.Dropdown(
                                            id="forn",
                                            options=[{"label": "[TODOS]", "value": "[TODOS]"}] + [{"label": x, "value": x} for x in lista_fornecedores0],
                                            value="[TODOS]",
                                            placeholder="Selecione...",
                                            clearable=False,
                                            className="shadow-sm"
                                        )
                                    ], lg=3, md=6, className="mb-3"),
                                    dbc.Col([
                                        html.Div("Fabricante", className="small text-muted fw-bold mb-1"),
                                        dcc.Dropdown(
                                            id="fab",
                                            options=[{"label": "[TODOS]", "value": "[TODOS]"}],
                                            value="[TODOS]",
                                            clearable=False,
                                            className="shadow-sm"
                                        )
                                    ], lg=3, md=6, className="mb-3"),
                                    dbc.Col([
                                        html.Div("Categoria", className="small text-muted fw-bold mb-1"),
                                        dcc.Dropdown(
                                            id="cat",
                                            options=[{"label": "[TODAS]", "value": "[TODAS]"}],
                                            value="[TODAS]",
                                            clearable=False,
                                            className="shadow-sm"
                                        )
                                    ], lg=3, md=6, className="mb-3"),
                                    dbc.Col([
                                        html.Div("Embalagem", className="small text-muted fw-bold mb-1"),
                                        dcc.Dropdown(
                                            id="tipo_embal",
                                            options=[{"label": x, "value": x} for x in _TIPO_EMBAL_OPCOES],
                                            value="[TODAS]",
                                            clearable=False,
                                            className="shadow-sm"
                                        )
                                    ], lg=3, md=6, className="mb-3"),
                                    dbc.Col([
                                        html.Div("Sim. Margem (%)", className="small text-dark fw-bold mb-1"),
                                        dbc.Input(
                                            id="meta_t1",
                                            value=30.0,
                                            type="number",
                                            step=0.1,
                                            className="shadow-sm text-dark fw-bold"
                                        )
                                    ], lg=2, md=6, className="mb-3"),
                                    dbc.Col([
                                        html.Div("Delta Alvo (%)", className="small text-dark fw-bold mb-1"),
                                        dbc.Input(
                                            id="meta_t2",
                                            value=0.0,
                                            type="number",
                                            step=0.1,
                                            className="shadow-sm text-dark fw-bold"
                                        )
                                    ], lg=2, md=6, className="mb-3"),
                                ], className="g-3 align-items-end"),
                            ]),
                            html.Div(
                                "",
                                className="small-muted",
                                style={"marginTop": "8px"},
                            ),
                        ]
                    )
                ),
                width=12,
            ),
            className="mb-2",
        ),

        dbc.Tabs(
            id="tabs",
            active_tab="tab-1",
            children=[
                dbc.Tab(
                    label="1. Visão de Custo",
                    tab_id="tab-1",
                    children=[
                        dcc.Loading(
                            type="circle",
                            children=html.Div(
                                [
                                    make_summary_block("t1"),
                                    html.Div("Objetivo: Definir Preço/Margem para calcular Custo Alvo.", className="small-muted"),
                                    make_grid("grid-t1", coldefs_t1),
                                ]
                            ),
                        )
                    ],
                ),
                dbc.Tab(
                    label="2. Visão de Precificação",
                    tab_id="tab-2",
                    children=[
                        dcc.Loading(
                            type="circle",
                            children=html.Div(
                                [
                                    make_summary_block("t2"),
                                    html.Div("Objetivo: Definir Delta % vs Concorrente para ver a Margem Resultante.", className="small-muted"),
                                    make_grid("grid-t2", coldefs_t2),
                                ]
                            ),
                        )
                    ],
                ),
                dbc.Tab(
                    label="3. Visão Categ (Fornecedores)",
                    tab_id="tab-3",
                    children=[
                        dbc.Card(
                            dbc.CardBody(
                                [
                                    html.Div(
                                        [
                                            html.Span("1. Selecione a Categoria (Principal): ", style={"fontWeight": "700", "color": "navy"}),
                                            dcc.Dropdown(
                                                id="cat_t3",
                                                options=[{"label": x, "value": x} for x in lista_categorias0],
                                                value=None,
                                                placeholder="Selecione...",
                                                style={"width": "260px", "display": "inline-block"},
                                                clearable=True,
                                            ),
                                            html.Span("  -->  ", style={"fontWeight": "700"}),
                                            html.Span("2. Fornecedor (Opcional): ", style={"fontWeight": "700", "color": "navy"}),
                                            dcc.Dropdown(
                                                id="forn_t3",
                                                options=[{"label": "[TODOS]", "value": "[TODOS]"}],
                                                value="[TODOS]",
                                                style={"width": "260px", "display": "inline-block"},
                                                clearable=False,
                                            ),
                                        ],
                                        style={"display": "flex", "gap": "10px", "alignItems": "center", "flexWrap": "wrap"},
                                    )
                                ]
                            ),
                            className="mb-2",
                            style={"backgroundColor": "#e1e1e1"},
                        ),
                        dcc.Loading(
                            type="circle",
                            children=html.Div(
                                [
                                    make_summary_block("t3"),
                                    make_grid("grid-t3", coldefs_t3),
                                ]
                            ),
                        ),
                    ],
                ),
            ],
        ),

        # ---------- Modal Financeiro (Tab 1) ----------
        dbc.Modal(
            id="modal-fin",
            is_open=False,
            size="lg",
            children=[
                dbc.ModalHeader(dbc.ModalTitle(id="modal-fin-title", children="Simulação Fin.")),
                dbc.ModalBody(
                    [
                        dbc.Tabs(
                            id="modal-fin-tabs",
                            active_tab="tab-marg",
                            children=[
                                dbc.Tab(
                                    label="Definir Margem",
                                    tab_id="tab-marg",
                                    children=[
                                        dbc.Row(
                                            [
                                                dbc.Col([dbc.Label("Preço R$:"), dcc.Input(id="fin-p", type="text", value="", style={"width": "120px", "textAlign": "right"})], md=6),
                                                dbc.Col([dbc.Label("Margem %:"), dcc.Input(id="fin-m", type="text", value="", style={"width": "120px", "textAlign": "right"})], md=6),
                                            ],
                                            className="g-2",
                                        )
                                    ],
                                ),
                                dbc.Tab(
                                    label="Definir Custo",
                                    tab_id="tab-cust",
                                    children=[
                                        dbc.Row(
                                            [
                                                dbc.Col([dbc.Label("Preço R$:"), dcc.Input(id="fin-p2", type="text", value="", style={"width": "120px", "textAlign": "right"})], md=6),
                                                dbc.Col([dbc.Label("Custo R$:"), dcc.Input(id="fin-c", type="text", value="", style={"width": "120px", "textAlign": "right"})], md=6),
                                            ],
                                            className="g-2",
                                        )
                                    ],
                                ),
                            ],
                        ),
                        html.Hr(),
                        html.Div("Salvar aplica simulação manual (equivalente ao desktop). Reset desativa a simulação manual.", className="small-muted"),
                    ]
                ),
                dbc.ModalFooter(
                    [
                        dbc.Button("Salvar", id="fin-save", color="primary", className="me-2"),
                        dbc.Button("Reset", id="fin-reset", color="secondary"),
                        dbc.Button("Fechar", id="fin-close", color="light", className="ms-auto"),
                    ]
                ),
            ],
        ),

        # ---------- Modal Mercado (Tab 2) ----------
        dbc.Modal(
            id="modal-mkt",
            is_open=False,
            size="lg",
            children=[
                dbc.ModalHeader(dbc.ModalTitle(id="modal-mkt-title", children="Simulação Mercado")) ,
                dbc.ModalBody(
                    [
                        html.Div(id="mkt-menor-conc", style={"color": "blue", "fontWeight": "700"}),
                        html.Hr(),
                        dbc.Row(
                            [
                                dbc.Col(
                                    [
                                        dbc.Label("Diferença Alvo (%):"),
                                        dcc.Input(id="mkt-delta", type="text", value="", style={"width": "140px", "textAlign": "right"}),
                                        html.Div("% (Ex: -5.0 para 5% abaixo)", className="small-muted"),
                                    ],
                                    md=6,
                                ),
                                dbc.Col(
                                    [
                                        dbc.Label("Preço Resultante Estimado:"),
                                        html.Div(id="mkt-preco-est", style={"fontStyle": "italic", "color": "gray"}),
                                    ],
                                    md=6,
                                ),
                            ],
                            className="g-2",
                        ),
                    ]
                ),
                dbc.ModalFooter(
                    [
                        dbc.Button("Salvar Delta", id="mkt-save", color="primary", className="me-2"),
                        dbc.Button("Usar Padrão", id="mkt-reset", color="secondary"),
                        dbc.Button("Fechar", id="mkt-close", color="light", className="ms-auto"),
                    ]
                ),
            ],
        ),
    ],
)


# =============================================================================
# Callbacks: atualizar listas ao mudar Mês/Ano
# =============================================================================
# Atualiza opções do seletor Mês/Ano a cada TTL (garante novos meses apareçam)
# =============================================================================
@app.callback(
    Output("col-mes-inicio", "style"),
    Output("col-mes-fim", "style"),
    Output("col-mes-ref", "style"),
    Input("periodo_tipo", "value"),
)
def toggle_mes_inicio(periodo_tipo):
    if periodo_tipo == "personalizado":
        return {"display": "block"}, {"display": "block"}, {"display": "none"}
    return {"display": "none"}, {"display": "none"}, {"display": "block"}


@app.callback(
    Output("mes_ref", "options"),
    Input("interval-refresh", "n_intervals"),
)
def refresh_mes_ref_options(_):
    _, _, _, _, _, _, month_ctx = _get_data_for_mes_ref(None, force_reload=True)
    available_safe = month_ctx.get("available_labels_safe") or []
    available_human = month_ctx.get("available_labels_human") or []
    if available_safe and available_human and len(available_safe) == len(available_human):
        opts = [{"label": h, "value": s} for s, h in zip(available_safe, available_human)]
        return opts
    return no_update


# =============================================================================
@app.callback(
    Output("forn", "options"),
    Output("forn", "value"),
    Output("cat_t3", "options"),
    Output("cat_t3", "value"),
    Input("mes_ref", "value"),
)
def on_mes_ref_change(mes_ref):
    df_base, _, _, _, lista_fornecedores, lista_categorias, _ = _get_data_for_mes_ref(mes_ref)

    forn_list = (lista_fornecedores or ["SEM DADOS"])
    forn_opts = [{"label": "[TODOS]", "value": "[TODOS]"}] + [{"label": x, "value": x} for x in forn_list]
    forn_val = "[TODOS]"

    cat_opts = [{"label": x, "value": x} for x in (lista_categorias or [])]
    return forn_opts, forn_val, cat_opts, None


def _set_header(coldefs, field, header):
    out = []
    for c in coldefs:
        c2 = dict(c)
        if c2.get("field") == field:
            c2["headerName"] = header
        out.append(c2)
    return out


@app.callback(
    Output("grid-t1", "columnDefs"),
    Output("grid-t2", "columnDefs"),
    Output("grid-t3", "columnDefs"),
    Input("btn-atualizar", "n_clicks"),
    State("mes_ref", "value"),
    State("periodo_tipo", "value"),
    State("mes_inicio", "value"),
    State("mes_fim", "value"),
)
def update_grid_headers(n_clicks_atualizar, mes_ref, periodo_tipo, mes_inicio, mes_fim):
    mes_inicio = _parse_ddmmyyyy_to_safe(mes_inicio)
    mes_fim = _parse_ddmmyyyy_to_safe(mes_fim)
    available = _get_available_safe()
    ref_efetivo = _resolve_mes_ref(periodo_tipo, mes_ref, mes_fim)
    mes_ini = _resolve_mes_inicio(periodo_tipo, ref_efetivo, mes_inicio, available)
    _, _, _, _, _, _, month_ctx = _get_data_for_mes_ref(ref_efetivo, mes_inicio_safe=mes_ini)
    lab = _closed_month_label(month_ctx)

    t1 = _set_header(coldefs_t1, "Qtd Ref", f"Qtd {lab}")
    t2 = _set_header(coldefs_t2, "Qtd Ref", f"Qtd {lab}")

    t3 = coldefs_t3
    t3 = _set_header(t3, "Fat Ref", f"Fat {lab}")
    t3 = _set_header(t3, "Margem Ref R$", f"R$ Margem {lab}")
    t3 = _set_header(t3, "Margem Ref %", f"% Margem {lab}")

    return t1, t2, t3


@app.callback(
    Output("fab", "options"),
    Output("fab", "value"),
    Output("cat", "options"),
    Output("cat", "value"),
    Input("mes_ref", "value"),
    Input("forn", "value"),
)
def on_fornecedor_change(mes_ref, forn):
    df_base, _, _, _, _, _, _ = _get_data_for_mes_ref(mes_ref)
    fab_opts, cat_opts = _get_fab_cat_options_for_supplier(df_base, forn)
    return (
        [{"label": x, "value": x} for x in fab_opts],
        "[TODOS]",
        [{"label": x, "value": x} for x in cat_opts],
        "[TODAS]",
    )


@app.callback(
    Output("forn_t3", "options"),
    Output("forn_t3", "value"),
    Input("mes_ref", "value"),
    Input("cat_t3", "value"),
)
def on_cat_t3_change(mes_ref, cat_t3):
    df_base, _, _, _, _, _, _ = _get_data_for_mes_ref(mes_ref)
    opts = _get_supplier_options_for_category(df_base, cat_t3)
    return [{"label": x, "value": x} for x in opts], "[TODOS]"


# =============================================================================
# Callback principal: atualizar tabelas + resumo
# =============================================================================
@app.callback(
    Output("grid-t1", "rowData"),
    Output("grid-t2", "rowData"),
    Output("grid-t3", "rowData"),

    Output("kpi-line-t1", "children"),
    Output("breakdown-t1", "children"),

    Output("kpi-line-t2", "children"),
    Output("breakdown-t2", "children"),

    Output("kpi-line-t3", "children"),
    Output("breakdown-t3", "children"),

    Input("btn-atualizar", "n_clicks"),
    Input("tabs", "active_tab"),
    State("mes_ref", "value"),
    State("forn", "value"),
    State("fab", "value"),
    State("cat", "value"),
    State("tipo_embal", "value"),
    State("meta_t1", "value"),
    State("meta_t2", "value"),
    State("cat_t3", "value"),
    State("forn_t3", "value"),
    Input("store-sim", "data"),
    State("periodo_tipo", "value"),
    State("mes_inicio", "value"),
    State("mes_fim", "value"),
)
def refresh_all(n_clicks_atualizar, active_tab, mes_ref, forn, fab, cat, tipo_embal, meta_t1, meta_t2, cat_t3, forn_t3, sim_store, periodo_tipo, mes_inicio_val, mes_fim_val):
    mes_inicio_val = _parse_ddmmyyyy_to_safe(mes_inicio_val)
    mes_fim_val = _parse_ddmmyyyy_to_safe(mes_fim_val)
    ref_efetivo = _resolve_mes_ref(periodo_tipo, mes_ref, mes_fim_val)
    mes_ini = _resolve_mes_inicio(periodo_tipo, ref_efetivo, mes_inicio_val, _get_available_safe())
    df_base, bench_ano, _, _, _, _, month_ctx = _get_data_for_mes_ref(ref_efetivo, force_reload=False, mes_inicio_safe=mes_ini)

    meta_t1_atual = _safe_float_percent(meta_t1, 0.30)
    meta_t2_atual = _safe_float_percent(meta_t2, 0.00)

    df_view_12 = _filter_tab12(df_base, forn, fab, cat, tipo_embal or "[TODAS]")
    rows_t1 = build_tab1_rows(df_view_12, sim_store, meta_t1_atual)
    sum_t1 = compute_summary(df_view_12, bench_ano, month_ctx=month_ctx, rows=rows_t1)

    rows_t2 = build_tab2_rows(df_view_12, sim_store, meta_t2_atual)
    sum_t2 = compute_summary(df_view_12, bench_ano, month_ctx=month_ctx, rows=rows_t2)

    df_view_3 = _filter_tab3(df_base, cat_t3, forn_t3)
    rows_t3 = build_tab3_rows(df_view_3)
    sum_t3 = compute_summary(df_view_3, bench_ano, month_ctx=month_ctx)

    def kpi_children(summary):
        return [
            _format_kpi("Fat. Total:", _format_currency_br(summary["fat_total"]), None),
            _format_kpi("Margem Média:", f"{summary['marg_pond']:.2%}", "blue"),
            html.Span("| ", style={"color": "#999"}),
            _format_kpi("Total SKUs:", str(summary["qtd_sku"]), None),
            _format_kpi("A:", str(summary["sku_a"]), "green"),
            _format_kpi("B:", str(summary["sku_b"]), "#bda404"),
            _format_kpi("C:", str(summary["sku_c"]), "red"),
        ]

    return (
        rows_t1,
        rows_t2,
        rows_t3,

        kpi_children(sum_t1),
        _breakdown_component(sum_t1["breakdown"], month_ctx),

        kpi_children(sum_t2),
        _breakdown_component(sum_t2["breakdown"], month_ctx),

        kpi_children(sum_t3),
        _breakdown_component(sum_t3["breakdown"], month_ctx),
    )


@app.callback(
    Output("grid-t1", "columnSize"),
    Output("grid-t2", "columnSize"),
    Output("grid-t3", "columnSize"),
    Input("tabs", "active_tab"),
)
def fit_columns_on_visible_tab(active_tab):
    if active_tab == "tab-1":
        return "sizeToFit", no_update, no_update
    if active_tab == "tab-2":
        return no_update, "sizeToFit", no_update
    if active_tab == "tab-3":
        return no_update, no_update, "sizeToFit"
    return no_update, no_update, no_update


# =============================================================================
# Atualizar histórico ao clicar em célula (tabs 1 e 2)
# =============================================================================
@app.callback(
    Output("hist-box-t1", "children"),
    Output("hist-box-t2", "children"),
    Input("grid-t1", "cellClicked"),
    Input("grid-t2", "cellClicked"),
    Input("grid-t1", "selectedRows"),
    Input("grid-t2", "selectedRows"),
    State("grid-t1", "rowData"),
    State("grid-t2", "rowData"),
    Input("mes_ref", "value"),
    Input("forn", "value"),
    Input("fab", "value"),
    Input("cat", "value"),
    Input("tabs", "active_tab"),
    State("periodo_tipo", "value"),
    State("mes_inicio", "value"),
    State("mes_fim", "value"),
)
def on_cell_click(cell1, cell2, sel1, sel2, rowData1, rowData2, mes_ref, forn, fab, cat, active_tab, periodo_tipo, mes_inicio_val, mes_fim_val):
    hist_default = {"produto": "Selecione...", "cod_barras": "-", "hist_6m": "-", "hist_3m": "-", "hist_ref": "-", "hist_pico": "-"}

    mes_inicio_val = _parse_ddmmyyyy_to_safe(mes_inicio_val)
    mes_fim_val = _parse_ddmmyyyy_to_safe(mes_fim_val)
    ref_efetivo = _resolve_mes_ref(periodo_tipo, mes_ref, mes_fim_val)
    mes_ini = _resolve_mes_inicio(periodo_tipo, ref_efetivo, mes_inicio_val, _get_available_safe())
    df_base, _, _, _, _, _, month_ctx = _get_data_for_mes_ref(ref_efetivo, mes_inicio_safe=mes_ini)

    trig = ctx.triggered_id

    if trig == "grid-t2" and active_tab == "tab-2" and isinstance(sel2, list) and len(sel2) > 0:
        r = sel2[0]
        produto_key = r.get("_produto_key") or r.get("id")
        if produto_key and produto_key in df_base.index:
            row = df_base.loc[produto_key]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            hist = build_history_payload(row, month_ctx)
            return no_update, _history_component(hist, "t2", month_ctx).children

    if trig == "grid-t1" and active_tab == "tab-1" and isinstance(sel1, list) and len(sel1) > 0:
        r = sel1[0]
        produto_key = r.get("_produto_key") or r.get("id")
        if produto_key and produto_key in df_base.index:
            row = df_base.loc[produto_key]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            hist = build_history_payload(row, month_ctx)
            return _history_component(hist, "t1", month_ctx).children, no_update
        
    if trig in ("mes_ref", "forn", "fab", "cat", "tabs"):
        empty_hist = {"produto": "-", "cod_barras": "-", "hist_6m": "-", "hist_3m": "-", "hist_ref": "-", "hist_pico": "-"}
        return (
            _history_component(empty_hist, "t1", month_ctx).children,
            _history_component(empty_hist, "t2", month_ctx).children,
        )

    if df_base is None or df_base.empty:
        return _history_component(hist_default, "t1").children, _history_component(hist_default, "t2").children

    if trig == "grid-t1" and active_tab == "tab-1" and isinstance(cell1, dict):
        row_grid = _row_from_event(cell1, rowData1)
        produto_key = (row_grid or {}).get("_produto_key") or (row_grid or {}).get("id")
        if produto_key and produto_key in df_base.index:
            row = df_base.loc[produto_key]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            return _history_component(build_history_payload(row, month_ctx), "t1", month_ctx).children, no_update

    if trig == "grid-t2" and active_tab == "tab-2" and isinstance(cell2, dict):
        row_grid = _row_from_event(cell2, rowData2)
        produto_key = (row_grid or {}).get("_produto_key") or (row_grid or {}).get("id")
        if produto_key and produto_key in df_base.index:
            row = df_base.loc[produto_key]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            return no_update, _history_component(build_history_payload(row, month_ctx), "t2", month_ctx).children

    return no_update, no_update


# =============================================================================
# Duplo clique: abrir modais
# =============================================================================
def _parse_float(val, default=0.0) -> float:
    try:
        if val is None:
            return default
        return float(str(val).replace(",", ".").strip())
    except Exception:
        return default


@app.callback(
    Output("modal-fin", "is_open"),
    Output("modal-fin-title", "children"),
    Output("fin-p", "value"),
    Output("fin-m", "value"),
    Output("fin-p2", "value"),
    Output("fin-c", "value"),

    Output("modal-mkt", "is_open"),
    Output("modal-mkt-title", "children"),
    Output("mkt-menor-conc", "children"),
    Output("mkt-delta", "value"),

    Output("store-selected", "data"),
    Output("store-sim", "data"),

    Input("grid-t1", "cellDoubleClicked"),
    Input("grid-t2", "cellDoubleClicked"),
    Input("fin-close", "n_clicks"),
    Input("mkt-close", "n_clicks"),

    Input("fin-save", "n_clicks"),
    Input("fin-reset", "n_clicks"),
    Input("mkt-save", "n_clicks"),
    Input("mkt-reset", "n_clicks"),

    State("tabs", "active_tab"),
    State("meta_t1", "value"),
    State("meta_t2", "value"),
    State("modal-fin", "is_open"),
    State("modal-mkt", "is_open"),

    State("modal-fin-tabs", "active_tab"),
    State("fin-p", "value"),
    State("fin-m", "value"),
    State("fin-p2", "value"),
    State("fin-c", "value"),
    State("mkt-delta", "value"),

    State("store-selected", "data"),
    State("store-sim", "data"),

    State("grid-t1", "rowData"),
    State("grid-t2", "rowData"),
    prevent_initial_call=True,
)
def modal_controller(
    cell1, cell2, fin_close, mkt_close,
    fin_save, fin_reset, mkt_save, mkt_reset,
    active_tab, meta_t1, meta_t2, fin_is_open, mkt_is_open,
    fin_tab, fin_p, fin_m, fin_p2, fin_c, mkt_delta,
    selected_state, sim_store, rowData1, rowData2
):
    trig = ctx.triggered_id

    fin_open = bool(fin_is_open)
    mkt_open = bool(mkt_is_open)

    sim = {
        "manual": dict((sim_store or {}).get("manual", {})),
        "conc": dict((sim_store or {}).get("conc", {})),
    }

    produto_key = (selected_state or {}).get("produto_key")
    area_sel = (selected_state or {}).get("area") or ""

    # ---------- FECHAR ----------
    if trig == "fin-close":
        return (
            False, no_update, no_update, no_update, no_update, no_update,
            mkt_open, no_update, no_update, no_update,
            no_update, no_update
        )

    if trig == "mkt-close":
        return (
            fin_open, no_update, no_update, no_update, no_update, no_update,
            False, no_update, no_update, no_update,
            no_update, no_update
        )

    # ---------- SALVAR / RESET FIN ----------
    if trig in ("fin-save", "fin-reset"):
        if not produto_key:
            return (
                fin_open, no_update, no_update, no_update, no_update, no_update,
                mkt_open, no_update, no_update, no_update,
                no_update, no_update
            )

        if trig == "fin-reset":
            sim["manual"].pop(produto_key, None)
            return (
                False, no_update, no_update, no_update, no_update, no_update,
                mkt_open, no_update, no_update, no_update,
                no_update, sim
            )

        # fin-save
        if fin_tab == "tab-marg":
            p = _parse_float(fin_p, 0.0)
            m_perc = _parse_float(fin_m, 0.0) / 100.0
            sim["manual"][produto_key] = {"ativa": True, "preco": p, "margem": m_perc}
        else:
            p = _parse_float(fin_p2, 0.0)
            c = _parse_float(fin_c, 0.0)
            m = float(calcular_margem_real_percentual(c, p, area=area_sel))
            sim["manual"][produto_key] = {"ativa": True, "preco": p, "margem": m}

        return (
            False, no_update, no_update, no_update, no_update, no_update,
            mkt_open, no_update, no_update, no_update,
            no_update, sim
        )

    # ---------- SALVAR / RESET MKT ----------
    if trig in ("mkt-save", "mkt-reset"):
        if not produto_key:
            return (
                fin_open, no_update, no_update, no_update, no_update, no_update,
                mkt_open, no_update, no_update, no_update,
                no_update, no_update
            )

        if trig == "mkt-reset":
            sim["conc"].pop(produto_key, None)
            return (
                fin_open, no_update, no_update, no_update, no_update, no_update,
                False, no_update, no_update, no_update,
                no_update, sim
            )

        d = _parse_float(mkt_delta, 0.0) / 100.0
        sim["conc"][produto_key] = {"ativa": True, "delta": d}

        return (
            fin_open, no_update, no_update, no_update, no_update, no_update,
            False, no_update, no_update, no_update,
            no_update, sim
        )

    # ---------- ABRIR (duplo clique) ----------
    meta_t1_atual = _safe_float_percent(meta_t1, 0.30)
    meta_t2_atual = _safe_float_percent(meta_t2, 0.00)

    if trig == "grid-t1" and active_tab == "tab-1" and isinstance(cell1, dict):
        row_grid = _row_from_event(cell1, rowData1)
        if not row_grid:
            return (
                fin_open, no_update, no_update, no_update, no_update, no_update,
                mkt_open, no_update, no_update, no_update,
                {"produto_key": None, "area": ""}, no_update
            )

        produto_key = row_grid.get("_produto_key") or row_grid.get("id")
        menor_conc = float(row_grid.get("_menor_conc", 0.0))
        p_atual = float(row_grid.get("_p_atual", 0.0))
        area = str(row_grid.get("_area") or row_grid.get("Categ") or "")

        sim_manual_ativa = bool(row_grid.get("_sim_manual_ativa", False))
        sim_preco_man = float(row_grid.get("_sim_preco_man", 0.0))
        sim_marg_man = float(row_grid.get("_sim_marg_man", 0.0))

        if sim_manual_ativa:
            val_p = sim_preco_man
            val_m = sim_marg_man * 100.0
        else:
            val_p = (menor_conc if menor_conc > 0 else p_atual)
            val_m = meta_t1_atual * 100.0

        custo_calc = float(calcular_custo_necessario(val_p, val_m / 100.0, area=area))
        fin_title = f"Financeiro: {str(row_grid.get('_produto_nome',''))[:30]}"

        return (
            True, fin_title, f"{val_p:.2f}", f"{val_m:.1f}", f"{val_p:.2f}", f"{custo_calc:.2f}",
            False, "Simulação Mercado", "", "",
            {"produto_key": produto_key, "area": area}, no_update
        )

    if trig == "grid-t2" and active_tab == "tab-2" and isinstance(cell2, dict):
        row_grid = _row_from_event(cell2, rowData2)
        if not row_grid:
            return (
                fin_open, no_update, no_update, no_update, no_update, no_update,
                mkt_open, no_update, no_update, no_update,
                no_update, no_update
            )

        produto_key = row_grid.get("_produto_key") or row_grid.get("id")
        menor_conc = float(row_grid.get("_menor_conc", 0.0))
        area = str(row_grid.get("_area") or row_grid.get("Categ") or "")

        sim_conc_ativa = bool(row_grid.get("_sim_conc_ativa", False))
        sim_conc_delta = float(row_grid.get("_sim_conc_delta", 0.0))

        delta_atual = sim_conc_delta if sim_conc_ativa else meta_t2_atual

        mkt_title = f"Mercado: {str(row_grid.get('_produto_nome',''))[:30]}"
        mkt_menor = f"Menor Concorrente: {_format_currency_br(menor_conc)}"
        mkt_delta_str = f"{delta_atual*100.0:.1f}"

        return (
            False, "Simulação Fin.", "", "", "", "",
            True, mkt_title, mkt_menor, mkt_delta_str,
            {"produto_key": produto_key, "area": area}, no_update
        )

    return (
        fin_open, no_update, no_update, no_update, no_update, no_update,
        mkt_open, no_update, no_update, no_update,
        no_update, no_update
    )


# ---------- Atualização do preço estimado no modal mercado (quando delta muda) ----------
@app.callback(
    Output("mkt-preco-est", "children"),
    Input("mkt-delta", "value"),
    Input("store-selected", "data"),
    Input("grid-t2", "rowData"),
)
def update_mkt_estimate(delta, selected, rowData):
    produto_key = (selected or {}).get("produto_key")
    if not produto_key or not rowData:
        return "-"

    try:
        d = float(str(delta).replace(",", ".")) / 100.0
    except Exception:
        return "-"

    row = next((r for r in rowData if r.get("_produto_key") == produto_key), None)
    if not row:
        return "-"

    menor_conc = float(row.get("_menor_conc", 0.0))
    p_atual = float(row.get("_p_atual", 0.0))
    p_est = (menor_conc * (1 + d)) if menor_conc > 0 else p_atual
    return _format_currency_br(p_est)


# ---------- Exportar ----------
@app.callback(
    Output("download-excel", "data"),
    Input("btn-export", "n_clicks"),
    State("mes_ref", "value"),
    State("tabs", "active_tab"),
    State("forn", "value"),
    State("fab", "value"),
    State("cat", "value"),
    State("cat_t3", "value"),
    State("forn_t3", "value"),
    State("store-sim", "data"),
    State("grid-t1", "columnState"),
    State("grid-t2", "columnState"),
    State("grid-t3", "columnState"),
    State("grid-t1", "rowData"),
    State("grid-t2", "rowData"),
    State("grid-t3", "rowData"),
    State("periodo_tipo", "value"),
    State("mes_inicio", "value"),
    State("mes_fim", "value"),
    prevent_initial_call=True,
)
def export_excel(_, mes_ref, active_tab, forn, fab, cat, cat_t3, forn_t3, sim_store, cs_t1, cs_t2, cs_t3, rd_t1, rd_t2, rd_t3, periodo_tipo, mes_inicio_val, mes_fim_val):
    try:
        mes_inicio_val = _parse_ddmmyyyy_to_safe(mes_inicio_val)
        mes_fim_val = _parse_ddmmyyyy_to_safe(mes_fim_val)
        ref_efetivo = _resolve_mes_ref(periodo_tipo, mes_ref, mes_fim_val)
        mes_ini = _resolve_mes_inicio(periodo_tipo, ref_efetivo, mes_inicio_val, _get_available_safe())
        df_base, _, _, _, _, _, month_ctx = _get_data_for_mes_ref(ref_efetivo, mes_inicio_safe=mes_ini)
        if df_base is None or df_base.empty:
            return no_update

        if active_tab == "tab-1":
            visible = _visible_fields_from_column_state(cs_t1)
            rows = rd_t1 or []
            nome_tipo = "FINANCEIRO"
        elif active_tab == "tab-2":
            visible = _visible_fields_from_column_state(cs_t2)
            rows = rd_t2 or []
            nome_tipo = "MERCADO"
        else:
            visible = _visible_fields_from_column_state(cs_t3)
            rows = rd_t3 or []
            nome_tipo = "CATEGORIA"

        if not rows:
            return no_update

        if not visible:
            if active_tab == "tab-1":
                visible = [c["field"] for c in coldefs_t1 if c.get("field")]
            elif active_tab == "tab-2":
                visible = [c["field"] for c in coldefs_t2 if c.get("field")]
            else:
                visible = [c["field"] for c in coldefs_t3 if c.get("field")]

        df_out = pd.DataFrame(rows)

        # remove colunas internas
        drop_tech = [c for c in df_out.columns if str(c).startswith("_") or str(c).startswith("__")]
        df_out.drop(columns=drop_tech, inplace=True, errors="ignore")

        # mantém só as visíveis
        keep = [c for c in visible if c in df_out.columns]
        if keep:
            df_out = df_out[keep]

        # converte strings formatadas para número real
        numeric_like_cols = {
            "Qtd Ref",
            "Faturamento",
            "Fat_Acum_Pct",
            "Preço Atual",
            "Custo",
            "Marg R$",
            "Marg %",
            "PETZ",
            "PROCAMPO",
            "Dif % (Menor)",
            "Sim Preço",
            "Sim Marg",
            "Sim Custo",
            "Marg Atual %",
            "Dif % (menor preço)",
            "DELTA ALVO %",
            "Sim Preço (Conc)",
            "Sim Margem (Result)",
            "Fat Ref",
            "Margem Ref R$",
            "Margem Ref %",
        }

        for col in df_out.columns:
            if col in numeric_like_cols:
                df_out[col] = df_out[col].map(_parse_br_number_like_excel)
                df_out[col] = pd.to_numeric(df_out[col], errors="coerce")

        mes_safe = (month_ctx or {}).get("ref_month_safe") or "MES"
        filename = f"Simulacao_{nome_tipo}_{mes_safe}.xlsx"

        tmpdir = Path(tempfile.gettempdir())
        tmp_path = tmpdir / f"dash_export_{nome_tipo}_{mes_safe}.xlsx"

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            df_out.to_excel(writer, index=False, sheet_name="Simulacao")

        wb = load_workbook(tmp_path)
        ws = wb["Simulacao"]
        _apply_excel_formats(ws)
        wb.save(tmp_path)

        return dcc.send_file(str(tmp_path), filename=filename)

    except Exception:
        logger.exception("Falha no export_excel()")
        return no_update


if __name__ == "__main__":
    import os
    debug = os.getenv("DEBUG", "false").lower() == "true"
    app.run(debug=debug, host="0.0.0.0", port=8050)
